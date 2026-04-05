"""Professional reports: Excel and PDF with letterhead, styled tables."""
from datetime import datetime
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Sum
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from core.permissions import get_user_role
from members.models import Member
from payments.models import Payment


def _parse_date_range(request):
    """Get start_date, end_date from query params or default to this month."""
    today = datetime.now().date()
    start = request.query_params.get('start_date')
    end = request.query_params.get('end_date')
    if start:
        try:
            start = datetime.strptime(start, '%Y-%m-%d').date()
        except ValueError:
            start = today.replace(day=1)
    else:
        start = today.replace(day=1)
    if end:
        try:
            end = datetime.strptime(end, '%Y-%m-%d').date()
        except ValueError:
            end = today
    else:
        end = today
    return start, end


def _staff_only(request):
    """Reports: only super_admin, admin, loan_officer (not member)."""
    role = get_user_role(request)
    return role in ('super_admin', 'admin', 'loan_officer')


# ---- Excel styling ----
def _excel_header_style(ws, row_num, cols, title=None):
    """Apply letterhead and header styling."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    fill_header = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    fill_sub = PatternFill(start_color='e8eef5', end_color='e8eef5', fill_type='solid')
    font_white = Font(bold=True, color='FFFFFF', size=11)
    font_dark = Font(bold=True, color='1e3a5f', size=10)
    thin_border = Border(
        left=Side(style='thin', color='c5c9cc'),
        right=Side(style='thin', color='c5c9cc'),
        top=Side(style='thin', color='c5c9cc'),
        bottom=Side(style='thin', color='c5c9cc'),
    )
    r = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1e3a5f')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 28
        r = 2
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill_header if r == row_num and not title else fill_sub
        cell.font = font_white if cell.fill == fill_header else font_dark
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row_num].height = 22


def _excel_data_style(ws, start_row, cols):
    """Apply borders and alignment to data rows."""
    from openpyxl.styles import Alignment, Border, Side
    thin = Side(style='thin', color='c5c9cc')
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            c.alignment = Alignment(vertical='center', wrap_text=True)


def _excel_footer(ws, cols, start_date, end_date):
    """Add footer with generated timestamp."""
    from openpyxl.styles import Font, Alignment
    r = ws.max_row + 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    ws.cell(row=r, column=1, value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Period: {start_date} to {end_date}")
    ws.cell(row=r, column=1).font = Font(size=9, color='6b7280', italic=True)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def report_members_excel(request):
    """Export members list to Excel with professional styling."""
    if not _staff_only(request):
        return Response({'detail': 'Permission denied. Staff roles only.'}, status=403)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    start, end = _parse_date_range(request)
    qs = Member.objects.filter(date_joined__gte=start, date_joined__lte=end).order_by('member_number')

    wb = Workbook()
    ws = wb.active
    ws.title = "Members Report"

    headers = ['Member No', 'First Name', 'Last Name', 'Email', 'Phone', 'Status', 'Membership Expiry', 'Date Joined']
    ws.append(["AHTTAK Membership System — Members Report"])
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1e3a5f')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append([f"Report Period: {start} to {end}", "", "", "", "", "", "", ""])
    ws.append([])
    ws.append(headers)

    header_row = 5
    _excel_header_style(ws, header_row, len(headers), title=None)
    # Override: we added title rows, so style the actual header row
    from openpyxl.styles import PatternFill
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        c.font = Font(bold=True, color='FFFFFF', size=10)

    for m in qs:
        ws.append([
            m.member_number, m.first_name, m.last_name, m.email or '', m.phone or '',
            m.get_status_display(), str(m.membership_expiry or ''), str(m.date_joined),
        ])

    # Column widths
    widths = [14, 14, 14, 26, 16, 14, 18, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _excel_data_style(ws, header_row + 1, len(headers))
    _excel_footer(ws, len(headers), start, end)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="AHTTAK_Members_{start}_{end}.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def report_payments_excel(request):
    """Export payments to Excel with professional styling."""
    if not _staff_only(request):
        return Response({'detail': 'Permission denied. Staff roles only.'}, status=403)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    start, end = _parse_date_range(request)
    qs = Payment.objects.filter(payment_date__gte=start, payment_date__lte=end).select_related('member').order_by('-payment_date')

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments Report"

    headers = ['Receipt No', 'Date', 'Member No', 'Member Name', 'Type', 'Amount (KES)', 'Method', 'Transaction Code']
    ws.append(["AHTTAK Membership System — Payments Report"])
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1e3a5f')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append([f"Report Period: {start} to {end}"])
    ws.append([])
    ws.append(headers)

    header_row = 5
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        c.font = Font(bold=True, color='FFFFFF', size=10)

    for m in qs:
        ws.append([
            m.receipt_number, str(m.payment_date or ''), m.member.member_number,
            f"{m.member.first_name} {m.member.last_name}", m.get_payment_type_display(),
            float(m.amount), m.get_method_display(), m.transaction_code or '',
        ])

    widths = [14, 12, 12, 24, 22, 14, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _excel_data_style(ws, header_row + 1, len(headers))
    _excel_footer(ws, len(headers), start, end)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="AHTTAK_Payments_{start}_{end}.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def report_summary_pdf(request):
    """Summary report as PDF with professional layout."""
    if not _staff_only(request):
        return Response({'detail': 'Permission denied. Staff roles only.'}, status=403)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    start, end = _parse_date_range(request)
    members_count = Member.objects.filter(date_joined__gte=start, date_joined__lte=end).count()
    payments_qs = Payment.objects.filter(payment_date__gte=start, payment_date__lte=end)
    total_collections = payments_qs.aggregate(s=Sum('amount'))['s'] or 0
    payments_count = payments_qs.count()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='ReportTitle',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#1e3a5f'),
        spaceAfter=6,
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        name='ReportSub',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#6b7280'),
        spaceAfter=12,
        alignment=1,
    )
    heading_style = ParagraphStyle(
        name='ReportHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#1e3a5f'),
        spaceAfter=8,
    )

    elements = [
        Paragraph("AHTTAK Membership System", title_style),
        Paragraph("Summary Report", subtitle_style),
        Paragraph(f"Report Period: {start} to {end}", styles['Normal']),
        Spacer(1, 12 * mm),
        Paragraph("Overview", heading_style),
    ]

    data = [
        ['Metric', 'Value'],
        ['New Members', str(members_count)],
        ['Total Payments', str(payments_count)],
        ['Total Collections (KES)', f"{total_collections:,.2f}"],
    ]
    t = Table(data, colWidths=[80 * mm, 80 * mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 16 * mm))
    elements.append(Paragraph(
        f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        ParagraphStyle(name='Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#9ca3af'), alignment=1)
    ))
    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="AHTTAK_Summary_{start}_{end}.pdf"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def report_list(request):
    """List available reports (for UI). Staff only."""
    if not _staff_only(request):
        return Response({'detail': 'Permission denied. Staff roles only.'}, status=403)
    return Response({
        'reports': [
            {'id': 'members_excel', 'name': 'Members (Excel)', 'url': '/api/reports/members/excel/', 'params': ['start_date', 'end_date']},
            {'id': 'payments_excel', 'name': 'Payments (Excel)', 'url': '/api/reports/payments/excel/', 'params': ['start_date', 'end_date']},
            {'id': 'summary_pdf', 'name': 'Summary (PDF)', 'url': '/api/reports/summary/pdf/', 'params': ['start_date', 'end_date']},
        ]
    })
